import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

class ExcelHandler:
    @staticmethod
    def get_sheet_names(file_path):
        if not file_path or not os.path.exists(file_path):
            raise Exception("文件不存在")
        if not file_path.lower().endswith(('.xlsx', '.xls')):
            raise Exception("不是有效的Excel文件")
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets
        except Exception as e:
            raise Exception(f"读取Excel失败: {str(e) if str(e) else '文件损坏或格式不支持'}")
    
    @staticmethod
    def read_sheet(file_path, sheet_name, header=0):
        if not file_path or not os.path.exists(file_path):
            raise Exception("文件不存在")
        if not sheet_name:
            raise Exception("未指定Sheet名称")
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=header, dtype=str, keep_default_na=False)
            return df.fillna('')
        except Exception as e:
            raise Exception(f"读取Sheet失败: {str(e) if str(e) else 'Sheet不存在或内容为空'}")
    
    @staticmethod
    def export_compare_result(result_df, output_path, diff_columns):
        try:
            # 导出列完全是原左表列，去掉内部字段
            internal_cols = ['__status', '__diff_info']
            export_cols = [col for col in result_df.columns if col not in internal_cols]
            export_df = result_df[export_cols]
            status_list = result_df['__status'].tolist() if '__status' in result_df.columns else []
            diff_info_list = result_df['__diff_info'].tolist() if '__diff_info' in result_df.columns else []
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                export_df.to_excel(writer, index=False, sheet_name='比对结果')
            
            wb = load_workbook(output_path)
            ws = wb['比对结果']
            
            # 使用最基础的颜色，确保兼容
            green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            diff_red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            
            col_map = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
            
            for row_idx in range(2, ws.max_row + 1):
                row_pos = row_idx - 2
                status = status_list[row_pos] if row_pos < len(status_list) else ''
                diff_info = diff_info_list[row_pos] if row_pos < len(diff_info_list) else {}
                if not diff_info:
                    diff_info = {}
                
                row_fill = None
                if status == 'added':
                    row_fill = green_fill
                elif status == 'modified':
                    row_fill = yellow_fill
                elif status == 'deleted':
                    row_fill = red_fill
                
                # 先填充整行颜色
                for col_name, col_idx in col_map.items():
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if row_fill:
                        cell.fill = row_fill
                    
                    # 差异单元格标红
                    if col_name in diff_info:
                        cell.fill = diff_red_fill
                        left_val, right_val = diff_info[col_name]
                        comment = Comment(f"左表值: {left_val}\n右表值: {right_val}", "比对工具")
                        comment.width = 200
                        comment.height = 60
                        cell.comment = comment
                
                # 行级批注
                if status == 'deleted':
                    cell = ws.cell(row=row_idx, column=1)
                    comment = Comment("右表无此记录", "比对工具")
                    cell.comment = comment
                elif status == 'added':
                    cell = ws.cell(row=row_idx, column=1)
                    comment = Comment("右表新增记录", "比对工具")
                    cell.comment = comment
            
            wb.save(output_path)
            return True
        except Exception as e:
            raise Exception(f"导出失败: {str(e)}")